# cerbereall.py V1.0.0 - 10/01/2026 07:48:58
__pgm__ = "cerbereall.py"
__version__ = "1.0.0"
__date__ = "10/01/2026 07:48:58"

r"""
╔═══════════════════════════════════════════════════════════════════
║ cerbere.py - Gardien des Enfers Numériques v2.0
║ Système GÉNÉRALISÉ de validation, versionning et sauvegarde
║ 
║ v2.0.0 : Version généralisée multi-cibles
║          - 3 points d'entrée : cerberetests, cerbereprod, cerbereall
║          - Règles de dispatch externalisées (cerbere_config.csv)
║          - Versionning indépendant par cible (grandlivre/journal)
║          - Exécution distante de commandes Python (*.exec.csv)
║          - Sauvegardes indépendantes par cible
║
║ Points d'entrée :
║   c:/cerberetests/  → c:/cx/ (environnement tests)
║   c:/cerbereprod/   → c:/kitviewsearchV5/ (environnement prod)
║   c:/cerbereall/    → Les deux cibles simultanément
║
║ Fichiers de configuration par cible :
║   - cerbere_config.csv : définition cible + règles de dispatch
║   - grandlivre.csv : situation actuelle des versions
║   - journal.csv : historique complet des versions
║
║ Exécution distante :
║   - Déposer un fichier *.exec.csv avec des commandes python
║   - Récupérer le résultat dans *.exec.done.csv
╚═══════════════════════════════════════════════════════════════════
"""

import os
import sys
import time
import shutil
import csv
import re
import subprocess
import fnmatch
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

# Dépendance externe pour lecture xlsx
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ═══════════════════════════════════════════════════════════════════
# CONFIGURATION PAR DÉFAUT
# ═══════════════════════════════════════════════════════════════════

# Points d'entrée
CERBERE_TESTS_DIR = r"c:\cerberetests"
CERBERE_PROD_DIR = r"c:\cerbereprod"
CERBERE_ALL_DIR = r"c:\cerbereall"

# Backup SSD
SSD_BASE_DIR = r"g:\cxssd"

# Timing (en secondes)
CYCLE_WAIT = 30

# ═══════════════════════════════════════════════════════════════════
# CLASSES DE CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

@dataclass
class RegleDispatch:
    """Une règle de dispatch (pattern, extension, destination)"""
    pattern: str        # Glob pattern (*, data*, *pat*, etc.)
    extension: str      # Extension avec point (.csv, .py, etc.) ou * pour tout
    destination: str    # Sous-répertoire cible (vide = racine)
    commentaire: str    # Description de la règle
    
    def match(self, filename: str) -> bool:
        """Vérifie si le fichier correspond à cette règle"""
        name_lower = filename.lower()
        ext = os.path.splitext(filename)[1].lower()
        
        # Vérifier l'extension
        if self.extension != '*' and ext != self.extension.lower():
            return False
        
        # Vérifier le pattern sur le nom (sans extension)
        name_without_ext = os.path.splitext(name_lower)[0]
        pattern_lower = self.pattern.lower()
        
        # Cas spécial : * matche tout
        if pattern_lower == '*':
            return True
        
        # Utiliser fnmatch pour le pattern matching
        return fnmatch.fnmatch(name_without_ext, pattern_lower)


@dataclass
class CibleConfig:
    """Configuration d'une cible de déploiement"""
    nom: str                              # "tests" ou "prod"
    chemin: str                           # c:\cx ou c:\kitviewsearchV5
    sauve: str                            # Répertoire de sauvegarde
    archives: str                         # Répertoire d'archives
    actif: bool = True                    # Cible active ou non
    regles: list[RegleDispatch] = field(default_factory=list)
    
    # Chemins calculés
    grandlivre_file: str = ""
    journal_file: str = ""
    dict_file: str = ""
    
    def __post_init__(self):
        """Calcule les chemins dérivés"""
        # grandlivre et journal sont dans le dossier cerbere, pas dans la cible
        # Ils seront définis par le PointEntree
        pass


@dataclass
class PointEntree:
    """Point d'entrée de fichiers (cerberetests, cerbereprod, cerbereall)"""
    chemin: str                           # c:\cerberetests, etc.
    nom: str                              # "tests", "prod", "all"
    cibles: list[CibleConfig] = field(default_factory=list)
    
    @property
    def grandlivre_file(self) -> str:
        return os.path.join(self.chemin, "grandlivre.csv")
    
    @property
    def journal_file(self) -> str:
        return os.path.join(self.chemin, "journal.csv")
    
    @property
    def config_file(self) -> str:
        return os.path.join(self.chemin, "cerbere_config.csv")


# ═══════════════════════════════════════════════════════════════════
# VARIABLES GLOBALES
# ═══════════════════════════════════════════════════════════════════

POINTS_ENTREE: list[PointEntree] = []

# ═══════════════════════════════════════════════════════════════════
# FONCTIONS UTILITAIRES
# ═══════════════════════════════════════════════════════════════════

def log(msg: str, indent: int = 0):
    """Affiche un message avec indentation"""
    prefix = "   " * indent
    print(f"{prefix}{msg}")


def timestamp_now() -> str:
    """Retourne le timestamp actuel au format jj/mm/aaaa hh:mm:ss"""
    return datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def date_today() -> str:
    """Retourne la date du jour au format jjmmaaaa"""
    return datetime.now().strftime("%d%m%Y")


def afficher_logo():
    """Affiche le logo ASCII de Cerbère"""
    logo = r"""
    ║                                                           
    ║              /\_/\           /\_/\           /\_/\        
    ║             ( o.o )         ( o.o )         ( o.o )       
    ║              > ^ <           > ^ <           > ^ <        
    ║             /|   |\         /|   |\         /|   |\       
    ║            (_|   |_)       (_|   |_)       (_|   |_)      
    ║                  \___________/  \___________/             
    ║                     CERBERE.py V2.0.0                     
    ║              Gardien Multi-Cibles Généralisé              
    ║                                                           
    """
    print(logo)


def ensure_directory(path: str):
    """Crée un répertoire s'il n'existe pas"""
    if not os.path.exists(path):
        os.makedirs(path)
        log(f"✓ Répertoire créé : {path}", 1)


def detect_encoding(filepath: str) -> str:
    """Détecte l'encodage d'un fichier (utf-8-sig ou utf-8)"""
    try:
        with open(filepath, 'rb') as f:
            raw = f.read(4)
            if raw.startswith(b'\xef\xbb\xbf'):
                return 'utf-8-sig'
            else:
                return 'utf-8'
    except Exception:
        return 'unknown'


def detect_csv_separator(filepath: str) -> str:
    """Détecte le séparateur utilisé dans un fichier CSV"""
    try:
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    for line in f:
                        line = line.strip()
                        if not line or line.startswith('#'):
                            continue
                        count_semicolon = line.count(';')
                        count_comma = line.count(',')
                        count_tab = line.count('\t')
                        
                        if count_semicolon >= count_comma and count_semicolon >= count_tab:
                            return ';'
                        elif count_tab >= count_comma:
                            return '\t'
                        else:
                            return ','
                break
            except UnicodeDecodeError:
                continue
    except Exception:
        pass
    return ';'


def check_csv_format(filepath: str) -> tuple[bool, bool, str]:
    """Vérifie le format d'un fichier CSV"""
    encoding = detect_encoding(filepath)
    encoding_ok = (encoding == 'utf-8-sig')
    separator = detect_csv_separator(filepath)
    separator_ok = (separator == ';')
    return encoding_ok, separator_ok, encoding


# ═══════════════════════════════════════════════════════════════════
# CHARGEMENT DE LA CONFIGURATION
# ═══════════════════════════════════════════════════════════════════

def charger_config_cible(config_file: str) -> Optional[CibleConfig]:
    """
    Charge la configuration d'une cible depuis un fichier cerbere_config.csv.
    Retourne None si le fichier n'existe pas ou est invalide.
    """
    if not os.path.exists(config_file):
        log(f"⚠️  Fichier config non trouvé : {config_file}", 1)
        return None
    
    try:
        cible_params = {}
        regles = []
        
        with open(config_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f, delimiter=';')
            for row in reader:
                # Ignorer lignes vides et commentaires
                if not row or row[0].startswith('#'):
                    continue
                
                row_type = row[0].strip().lower()
                
                if row_type == 'cible' and len(row) >= 3:
                    param = row[1].strip().lower()
                    valeur = row[2].strip()
                    cible_params[param] = valeur
                    
                elif row_type == 'regle' and len(row) >= 4:
                    pattern = row[1].strip()
                    extension = row[2].strip()
                    destination = row[3].strip()
                    commentaire = row[4].strip() if len(row) > 4 else ""
                    regles.append(RegleDispatch(pattern, extension, destination, commentaire))
        
        # Créer la cible
        if 'nom' not in cible_params or 'chemin' not in cible_params:
            log(f"❌ Config incomplète (nom/chemin manquant) : {config_file}", 1)
            return None
        
        cible = CibleConfig(
            nom=cible_params.get('nom', ''),
            chemin=cible_params.get('chemin', ''),
            sauve=cible_params.get('sauve', ''),
            archives=cible_params.get('archives', ''),
            actif=cible_params.get('actif', 'oui').lower() in ['oui', 'yes', 'true', '1'],
            regles=regles
        )
        
        return cible
        
    except Exception as e:
        log(f"❌ Erreur lecture config : {config_file} - {e}", 1)
        return None


def charger_configuration() -> list[PointEntree]:
    """
    Charge la configuration de tous les points d'entrée.
    Retourne la liste des PointEntree configurés.
    """
    points = []
    
    # Point d'entrée TESTS
    if os.path.exists(CERBERE_TESTS_DIR):
        pe_tests = PointEntree(chemin=CERBERE_TESTS_DIR, nom="tests")
        cible = charger_config_cible(pe_tests.config_file)
        if cible and cible.actif:
            pe_tests.cibles.append(cible)
            points.append(pe_tests)
            log(f"✓ Point d'entrée TESTS : {CERBERE_TESTS_DIR} → {cible.chemin}", 1)
    
    # Point d'entrée PROD
    if os.path.exists(CERBERE_PROD_DIR):
        pe_prod = PointEntree(chemin=CERBERE_PROD_DIR, nom="prod")
        cible = charger_config_cible(pe_prod.config_file)
        if cible and cible.actif:
            pe_prod.cibles.append(cible)
            points.append(pe_prod)
            log(f"✓ Point d'entrée PROD : {CERBERE_PROD_DIR} → {cible.chemin}", 1)
    
    # Point d'entrée ALL (combine les cibles des deux autres)
    if os.path.exists(CERBERE_ALL_DIR):
        pe_all = PointEntree(chemin=CERBERE_ALL_DIR, nom="all")
        # Récupérer les cibles depuis tests et prod
        for pe in points:
            for cible in pe.cibles:
                pe_all.cibles.append(cible)
        if pe_all.cibles:
            points.append(pe_all)
            cibles_noms = [c.nom for c in pe_all.cibles]
            log(f"✓ Point d'entrée ALL : {CERBERE_ALL_DIR} → {', '.join(cibles_noms)}", 1)
    
    return points


# ═══════════════════════════════════════════════════════════════════
# CONVERSION DES FICHIERS
# ═══════════════════════════════════════════════════════════════════

def convert_csv_to_standard(filepath: str) -> bool:
    """
    Convertit un fichier CSV au format standard :
    - Encodage : UTF-8 avec BOM (utf-8-sig)
    - Séparateur : point-virgule (;)
    """
    filename = os.path.basename(filepath)
    log(f"🔄 Conversion CSV : {filename}", 1)
    
    try:
        current_separator = detect_csv_separator(filepath)
        content_lines = []
        detected_encoding = None
        
        for encoding in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                with open(filepath, 'r', encoding=encoding) as f:
                    content_lines = f.readlines()
                detected_encoding = encoding
                break
            except UnicodeDecodeError:
                continue
        
        if not content_lines:
            log(f"❌ Impossible de lire le fichier avec les encodages connus", 2)
            return False
        
        log(f"   Encodage détecté : {detected_encoding}", 2)
        log(f"   Séparateur détecté : '{current_separator}'", 2)
        
        if current_separator != ';':
            new_lines = []
            for line in content_lines:
                if line.strip().startswith('#'):
                    new_lines.append(line)
                else:
                    new_lines.append(line.replace(current_separator, ';'))
            content_lines = new_lines
            log(f"   Séparateur converti : '{current_separator}' → ';'", 2)
        
        with open(filepath, 'w', encoding='utf-8-sig', newline='') as f:
            f.writelines(content_lines)
        
        log(f"✓ Conversion réussie : {filename}", 1)
        return True
        
    except Exception as e:
        log(f"❌ Erreur conversion CSV : {e}", 2)
        return False


def convert_xlsx_to_csv(xlsx_path: str, csv_path: str) -> bool:
    """Convertit un fichier xlsx en csv (utf-8-sig, séparateur ;)"""
    if not OPENPYXL_AVAILABLE:
        log("❌ openpyxl non installé - conversion xlsx impossible", 1)
        return False
    
    xlsx_filename = os.path.basename(xlsx_path)
    csv_filename = os.path.basename(csv_path)
    
    try:
        log(f"📊 Conversion xlsx → csv : {xlsx_filename}", 1)
        
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        
        with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f, delimiter=';')
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell) if cell is not None else '' for cell in row]
                writer.writerow(row_data)
        
        wb.close()
        log(f"✓ Converti : {xlsx_filename} → {csv_filename}", 1)
        return True
        
    except Exception as e:
        log(f"❌ Erreur conversion xlsx : {e}", 2)
        return False


def traiter_xlsx_preliminaire(filepath: str, cible: CibleConfig) -> bool:
    """
    Traitement préliminaire d'un fichier xlsx :
    Vérifie si un CSV de même nom existe dans refs de la cible.
    """
    filename = os.path.basename(filepath)
    name_without_ext = os.path.splitext(filename)[0]
    csv_filename = name_without_ext + '.csv'
    refs_dir = os.path.join(cible.chemin, 'refs')
    csv_path_in_refs = os.path.join(refs_dir, csv_filename)
    
    if os.path.exists(csv_path_in_refs):
        log(f"📋 xlsx avec CSV correspondant trouvé : {csv_filename} (cible: {cible.nom})", 1)
        if convert_xlsx_to_csv(filepath, csv_path_in_refs):
            log(f"✓ CSV mis à jour dans refs/ : {csv_filename}", 1)
            return True
    
    return False


# ═══════════════════════════════════════════════════════════════════
# VALIDATION DES FICHIERS
# ═══════════════════════════════════════════════════════════════════

def valider_fichier(filepath: str) -> tuple[bool, str]:
    """Valide un fichier et tente la conversion automatique si CSV mal formaté"""
    filename = os.path.basename(filepath)
    ext = os.path.splitext(filename)[1].lower()
    
    if ext == '.csv':
        encoding_ok, separator_ok, detected_encoding = check_csv_format(filepath)
        
        if not encoding_ok or not separator_ok:
            log(f"⚠️  CSV non standard détecté : {filename}", 1)
            if not encoding_ok:
                log(f"   Encodage : {detected_encoding} (attendu utf-8-sig)", 2)
            if not separator_ok:
                log(f"   Séparateur : non standard (attendu ;)", 2)
            
            if convert_csv_to_standard(filepath):
                return True, ""
            else:
                return False, "Échec de la conversion automatique du CSV"
        
        return True, ""
    else:
        return True, ""


# ═══════════════════════════════════════════════════════════════════
# SYSTÈME DE VERSIONNING
# ═══════════════════════════════════════════════════════════════════

def init_version_files(point_entree: PointEntree):
    """Initialise les fichiers de version pour un point d'entrée"""
    # Pas de fichiers de version pour "all" (utilise ceux des cibles)
    if point_entree.nom == "all":
        return
    
    header = "nomdufichier;version;timestamp\n"
    
    if not os.path.exists(point_entree.grandlivre_file):
        with open(point_entree.grandlivre_file, 'w', encoding='utf-8-sig') as f:
            f.write(header)
        log(f"✓ Fichier grandlivre.csv initialisé ({point_entree.nom})", 1)
    
    if not os.path.exists(point_entree.journal_file):
        with open(point_entree.journal_file, 'w', encoding='utf-8-sig') as f:
            f.write(header)
        log(f"✓ Fichier journal.csv initialisé ({point_entree.nom})", 1)


def load_grandlivre(grandlivre_file: str) -> dict:
    """Charge le grand livre depuis un fichier"""
    result = {}
    if not os.path.exists(grandlivre_file):
        return result
    try:
        with open(grandlivre_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                if row.get('nomdufichier'):
                    result[row['nomdufichier']] = (row.get('version', 'V1.0.0'), row.get('timestamp', ''))
    except Exception as e:
        log(f"ERREUR lecture grandlivre : {e}")
    return result


def save_grandlivre(grandlivre_file: str, data: dict):
    """Sauvegarde le grand livre"""
    try:
        with open(grandlivre_file, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(['nomdufichier', 'version', 'timestamp'])
            for filename, (version, ts) in sorted(data.items()):
                writer.writerow([filename, version, ts])
    except Exception as e:
        log(f"ERREUR écriture grandlivre : {e}")


def append_journal(journal_file: str, filename: str, version: str, timestamp: str):
    """Ajoute une entrée au journal"""
    try:
        with open(journal_file, 'a', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow([filename, version, timestamp])
    except Exception as e:
        log(f"ERREUR écriture journal : {e}")


def increment_version(version: str) -> str:
    """Incrémente le numéro de version (PATCH)"""
    match = re.match(r'([vV])(\d+)\.(\d+)\.(\d+)', version)
    if match:
        prefix = match.group(1)
        major = match.group(2)
        minor = match.group(3)
        patch = int(match.group(4)) + 1
        return f"{prefix}{major}.{minor}.{patch}"
    return version


def get_next_version(filename: str, point_entree: PointEntree) -> str:
    """Détermine la prochaine version d'un fichier"""
    grandlivre_file = point_entree.grandlivre_file
    journal_file = point_entree.journal_file
    
    grandlivre = load_grandlivre(grandlivre_file)
    ts = timestamp_now()
    
    if filename in grandlivre:
        old_version, old_ts = grandlivre[filename]
        new_version = increment_version(old_version)
        log(f"📝 {filename} : {old_version} → {new_version} ({point_entree.nom})", 1)
    else:
        new_version = "V1.0.0"
        log(f"📝 {filename} : {new_version} (nouveau) ({point_entree.nom})", 1)
    
    grandlivre[filename] = (new_version, ts)
    save_grandlivre(grandlivre_file, grandlivre)
    append_journal(journal_file, filename, new_version, ts)
    
    return new_version


# ═══════════════════════════════════════════════════════════════════
# MISE À JOUR DES VERSIONS DANS LES FICHIERS
# ═══════════════════════════════════════════════════════════════════

def update_version_py(filepath: str, version: str, timestamp: str):
    """Met à jour la version dans un fichier Python"""
    filename = os.path.basename(filepath)
    version_num = version.lstrip('Vv')
    comment_line = f"# {filename} {version} - {timestamp}\n"
    pgm_line = f'__pgm__ = "{filename}"\n'
    version_line = f'__version__ = "{version_num}"\n'
    date_line = f'__date__ = "{timestamp}"\n'
    
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        lines = [l for l in lines if l.strip() != '#*TO*#']
        
        comment_found = pgm_found = version_found = date_found = False
        for i in range(min(20, len(lines))):
            line = lines[i]
            if line.startswith(f"# {filename}"):
                lines[i] = comment_line
                comment_found = True
            elif '__pgm__' in line:
                lines[i] = pgm_line
                pgm_found = True
            elif '__version__' in line:
                lines[i] = version_line
                version_found = True
            elif '__date__' in line:
                lines[i] = date_line
                date_found = True
        
        if not comment_found: lines.insert(0, comment_line)
        if not pgm_found: lines.insert(1, pgm_line)
        if not version_found: lines.insert(2, version_line)
        if not date_found: lines.insert(3, date_line)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.writelines(lines)
    except Exception as e:
        log(f"ERREUR mise à jour version .py : {e}", 1)


def update_version_csv(filepath: str, version: str, timestamp: str):
    """Met à jour la version dans un fichier CSV"""
    filename = os.path.basename(filepath)
    comment_line = f"# {filename} {version} - {timestamp}\n"
    try:
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            lines = f.readlines()
        if lines and lines[0].startswith('#'):
            lines[0] = comment_line
        else:
            lines.insert(0, comment_line)
        with open(filepath, 'w', encoding='utf-8-sig') as f:
            f.writelines(lines)
    except Exception as e:
        log(f"ERREUR mise à jour version .csv : {e}", 1)


def update_version_md(filepath: str, version: str, timestamp: str):
    """Met à jour la version dans un fichier Markdown"""
    filename = os.path.basename(filepath)
    basename = os.path.splitext(filename)[0]
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        prompt_found = any('Prompt' in line or 'prompt' in line for line in lines[:10])
        if not prompt_found:
            title_line = f"# Prompt {basename} {version} - {timestamp}\n\n"
            lines.insert(0, title_line)
            with open(filepath, 'w', encoding='utf-8') as f:
                f.writelines(lines)
    except Exception as e:
        log(f"ERREUR mise à jour version .md : {e}", 1)


def update_version_html(filepath: str, version: str, timestamp: str):
    """Met à jour la version dans un fichier HTML"""
    filename = os.path.basename(filepath)
    comment_line = f" * {filename} {version} - {timestamp}\n"
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        if '/*' in content:
            pattern = r'(/\*[\s\S]*?\*/)'
            match = re.search(pattern, content)
            if match:
                new_comment = f"/*\n{comment_line} */"
                content = content.replace(match.group(1), new_comment, 1)
            else:
                content = f"/*\n{comment_line} */\n" + content
        else:
            content = f"/*\n{comment_line} */\n" + content
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(content)
    except Exception as e:
        log(f"ERREUR mise à jour version .html : {e}", 1)


def update_file_version(filepath: str, version: str, timestamp: str):
    """Met à jour la version selon l'extension"""
    ext = os.path.splitext(filepath)[1].lower()
    
    if ext == '.py':
        update_version_py(filepath, version, timestamp)
    elif ext == '.csv':
        update_version_csv(filepath, version, timestamp)
    elif ext == '.md':
        update_version_md(filepath, version, timestamp)
    elif ext == '.html':
        update_version_html(filepath, version, timestamp)
    # .txt, .db, .xlsx, etc. : PAS DE MODIFICATION


# ═══════════════════════════════════════════════════════════════════
# DISPATCHING DES FICHIERS
# ═══════════════════════════════════════════════════════════════════

def get_destination_dir(filename: str, cible: CibleConfig) -> str:
    """
    Détermine le répertoire de destination selon les règles de la cible.
    Parcourt les règles dans l'ordre et retourne dès qu'une règle matche.
    """
    for regle in cible.regles:
        if regle.match(filename):
            if regle.destination:
                return os.path.join(cible.chemin, regle.destination)
            else:
                return cible.chemin
    
    # Par défaut : ztri
    return os.path.join(cible.chemin, 'ztri')


def dispatch_file_to_cible(filepath: str, cible: CibleConfig, version: str, timestamp: str) -> bool:
    """Dispatche un fichier vers une cible spécifique"""
    filename = os.path.basename(filepath)
    dest_dir = get_destination_dir(filename, cible)
    dest_path = os.path.join(dest_dir, filename)
    
    try:
        ensure_directory(dest_dir)
        shutil.copy2(filepath, dest_path)
        log(f"✓ {filename} → {dest_path}", 1)
        return True
    except Exception as e:
        log(f"ERREUR dispatch {filename} vers {cible.nom} : {e}", 1)
        return False


# ═══════════════════════════════════════════════════════════════════
# EXÉCUTION DISTANTE DE COMMANDES
# ═══════════════════════════════════════════════════════════════════

def executer_commandes(exec_file: str, cible: CibleConfig) -> str:
    """
    Exécute les commandes d'un fichier .exec.csv.
    Retourne le chemin du fichier .exec.done.csv créé.
    """
    log(f"⚡ Exécution distante : {os.path.basename(exec_file)} (cible: {cible.nom})", 1)
    
    commandes = []
    resultats = []
    
    try:
        # Lire les commandes
        with open(exec_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                if row.get('commande'):
                    commandes.append({
                        'commande': row.get('commande', '').strip(),
                        'arguments': row.get('arguments', '').strip()
                    })
        
        if not commandes:
            log(f"   Aucune commande trouvée", 2)
            return ""
        
        log(f"   {len(commandes)} commande(s) à exécuter", 2)
        
        # Exécuter chaque commande
        for cmd in commandes:
            commande = cmd['commande']
            arguments = cmd['arguments']
            
            # Sécurité : seules les commandes python sont autorisées
            if not commande.lower().startswith('python'):
                resultats.append({
                    **cmd,
                    'statut': 'REFUSÉ',
                    'resultat': 'Seules les commandes python sont autorisées',
                    'duree_sec': '0'
                })
                log(f"   ❌ Commande refusée : {commande}", 2)
                continue
            
            # Construire la commande complète
            full_cmd = f"{commande} {arguments}".strip()
            log(f"   🔄 {full_cmd}", 2)
            
            start_time = time.time()
            try:
                # Exécuter dans le répertoire de la cible
                result = subprocess.run(
                    full_cmd,
                    shell=True,
                    cwd=cible.chemin,
                    capture_output=True,
                    text=True,
                    timeout=300  # Timeout de 5 minutes
                )
                
                duree = time.time() - start_time
                
                if result.returncode == 0:
                    # Prendre les dernières lignes de stdout (max 500 chars)
                    output = result.stdout.strip()
                    if len(output) > 500:
                        output = "..." + output[-497:]
                    resultats.append({
                        **cmd,
                        'statut': 'OK',
                        'resultat': output.replace(';', ',').replace('\n', ' | '),
                        'duree_sec': f"{duree:.1f}"
                    })
                    log(f"   ✓ OK ({duree:.1f}s)", 2)
                else:
                    # Prendre les dernières lignes de stderr
                    error = result.stderr.strip()
                    if len(error) > 500:
                        error = "..." + error[-497:]
                    resultats.append({
                        **cmd,
                        'statut': 'ERREUR',
                        'resultat': error.replace(';', ',').replace('\n', ' | '),
                        'duree_sec': f"{duree:.1f}"
                    })
                    log(f"   ❌ ERREUR ({duree:.1f}s)", 2)
                    
            except subprocess.TimeoutExpired:
                duree = time.time() - start_time
                resultats.append({
                    **cmd,
                    'statut': 'TIMEOUT',
                    'resultat': 'Exécution interrompue après 5 minutes',
                    'duree_sec': f"{duree:.1f}"
                })
                log(f"   ⏰ TIMEOUT ({duree:.1f}s)", 2)
                
            except Exception as e:
                duree = time.time() - start_time
                resultats.append({
                    **cmd,
                    'statut': 'EXCEPTION',
                    'resultat': str(e).replace(';', ','),
                    'duree_sec': f"{duree:.1f}"
                })
                log(f"   💥 EXCEPTION : {e}", 2)
        
        # Créer le fichier de résultats
        done_file = exec_file.replace('.exec.csv', '.exec.done.csv')
        duree_totale = sum(float(r.get('duree_sec', 0)) for r in resultats)
        
        with open(done_file, 'w', encoding='utf-8-sig', newline='') as f:
            f.write(f"# Exécution terminée le {timestamp_now()}\n")
            f.write(f"# Cible : {cible.nom} ({cible.chemin})\n")
            f.write(f"# Durée totale : {duree_totale:.1f} secondes\n")
            f.write("#\n")
            
            writer = csv.DictWriter(f, fieldnames=['commande', 'arguments', 'statut', 'resultat', 'duree_sec'], delimiter=';')
            writer.writeheader()
            for r in resultats:
                writer.writerow(r)
        
        # Supprimer le fichier original
        os.remove(exec_file)
        
        log(f"✓ Résultats : {os.path.basename(done_file)}", 1)
        return done_file
        
    except Exception as e:
        log(f"❌ Erreur exécution : {e}", 1)
        return ""


def traiter_exec_files(point_entree: PointEntree):
    """Traite tous les fichiers .exec.csv d'un point d'entrée"""
    exec_files = [f for f in os.listdir(point_entree.chemin) 
                  if f.endswith('.exec.csv') and os.path.isfile(os.path.join(point_entree.chemin, f))]
    
    if not exec_files:
        return
    
    log(f"⚡ {len(exec_files)} fichier(s) d'exécution trouvé(s)", 1)
    
    for exec_filename in exec_files:
        exec_filepath = os.path.join(point_entree.chemin, exec_filename)
        
        for cible in point_entree.cibles:
            if cible.actif:
                executer_commandes(exec_filepath, cible)
                
                # Pour "all", le fichier est traité une fois mais exécuté sur chaque cible
                # On doit donc recopier le fichier pour la cible suivante
                if point_entree.nom == "all" and len(point_entree.cibles) > 1:
                    # Recréer le fichier pour la prochaine cible
                    # (il a été supprimé après la première exécution)
                    break  # En fait, on ne peut pas le refaire sans le contenu original
        
        # Pour "all", exécuter sur toutes les cibles
        if point_entree.nom == "all":
            # Lire d'abord le contenu
            try:
                with open(exec_filepath, 'r', encoding='utf-8-sig') as f:
                    content = f.read()
                
                for i, cible in enumerate(point_entree.cibles):
                    if cible.actif:
                        # Créer un fichier temporaire avec le suffixe de la cible
                        temp_file = exec_filepath.replace('.exec.csv', f'_{cible.nom}.exec.csv')
                        with open(temp_file, 'w', encoding='utf-8-sig') as f:
                            f.write(content)
                        executer_commandes(temp_file, cible)
                
                # Supprimer le fichier original
                if os.path.exists(exec_filepath):
                    os.remove(exec_filepath)
                    
            except Exception as e:
                log(f"❌ Erreur traitement exec ALL : {e}", 1)


# ═══════════════════════════════════════════════════════════════════
# IMPORT DES FICHIERS
# ═══════════════════════════════════════════════════════════════════

def traiter_fichier(filepath: str, point_entree: PointEntree) -> bool:
    """
    Traite un fichier entrant pour toutes les cibles du point d'entrée.
    Pour "all", le fichier va vers toutes les cibles.
    """
    filename = os.path.basename(filepath)
    
    # Ignorer les fichiers spéciaux
    if filename in ['grandlivre.csv', 'journal.csv', 'cerbere_config.csv']:
        return False
    if filename.endswith('.exec.csv') or filename.endswith('.exec.done.csv'):
        return False
    
    # Validation et conversion si nécessaire
    est_valide, msg_erreur = valider_fichier(filepath)
    if not est_valide:
        log(f"❌ ERREUR VALIDATION : {filename}", 1)
        log(f"   Raison : {msg_erreur}", 2)
        return False
    
    # Déterminer le point d'entrée pour le versionning
    # Pour "all", on utilise le grandlivre du premier cible trouvé (tests)
    if point_entree.nom == "all":
        # Trouver le point d'entrée tests pour le versionning
        for pe in POINTS_ENTREE:
            if pe.nom == "tests":
                version_pe = pe
                break
        else:
            version_pe = POINTS_ENTREE[0] if POINTS_ENTREE else point_entree
    else:
        version_pe = point_entree
    
    # Obtenir la version
    version = get_next_version(filename, version_pe)
    ts = timestamp_now()
    
    # Mettre à jour la version dans le fichier
    update_file_version(filepath, version, ts)
    
    # Dispatcher vers chaque cible
    success = True
    for cible in point_entree.cibles:
        if cible.actif:
            # Traitement préliminaire xlsx si applicable
            if filename.lower().endswith('.xlsx'):
                traiter_xlsx_preliminaire(filepath, cible)
            
            if not dispatch_file_to_cible(filepath, cible, version, ts):
                success = False
    
    # Supprimer le fichier source après dispatch
    if success and os.path.exists(filepath):
        os.remove(filepath)
    
    return success


def import_files(point_entree: PointEntree) -> int:
    """Importe les fichiers d'un point d'entrée"""
    if not os.path.exists(point_entree.chemin):
        ensure_directory(point_entree.chemin)
        return 0
    
    # Phase 0 : Exécution des commandes
    traiter_exec_files(point_entree)
    
    # Lister les fichiers (en excluant les fichiers spéciaux)
    all_files = os.listdir(point_entree.chemin)
    files = [f for f in all_files 
             if os.path.isfile(os.path.join(point_entree.chemin, f))
             and f not in ['grandlivre.csv', 'journal.csv', 'cerbere_config.csv']
             and not f.endswith('.exec.csv')
             and not f.endswith('.exec.done.csv')]
    
    if not files:
        log(f"✓ Aucun import ({point_entree.nom})", 1)
        return 0
    
    log(f"✓ {len(files)} fichier(s) détecté(s) ({point_entree.nom})", 1)
    
    # Phase 1 : Traitement préliminaire xlsx
    xlsx_files = [f for f in files if f.lower().endswith('.xlsx')]
    for filename in xlsx_files:
        filepath = os.path.join(point_entree.chemin, filename)
        for cible in point_entree.cibles:
            if cible.actif:
                traiter_xlsx_preliminaire(filepath, cible)
    
    # Phase 2 : Traitement normal
    success_count = 0
    for filename in files:
        filepath = os.path.join(point_entree.chemin, filename)
        if os.path.exists(filepath):  # Vérifier car xlsx pourrait avoir été déplacé
            if traiter_fichier(filepath, point_entree):
                success_count += 1
    
    return success_count


# ═══════════════════════════════════════════════════════════════════
# SAUVEGARDE
# ═══════════════════════════════════════════════════════════════════

def build_file_dict(root_dir: str) -> dict:
    """Construit un dictionnaire des fichiers avec leurs timestamps"""
    file_dict = {}
    if not os.path.exists(root_dir):
        return file_dict
    for root, _, files in os.walk(root_dir):
        for f in files:
            full_path = os.path.join(root, f)
            rel_path = os.path.relpath(full_path, root_dir)
            try:
                ts = os.path.getmtime(full_path)
                file_dict[rel_path] = ts
            except Exception:
                pass
    return file_dict


def save_dict_to_csv(file_dict: dict, path: str):
    """Sauvegarde le dictionnaire dans un fichier CSV"""
    try:
        with open(path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(['path', 'timestamp'])
            for k, v in file_dict.items():
                writer.writerow([k, v])
    except Exception as e:
        log(f"ERREUR sauvegarde dictionnaire : {e}", 1)


def load_dict_from_csv(path: str) -> dict:
    """Charge le dictionnaire depuis un fichier CSV"""
    if not os.path.exists(path):
        return {}
    result = {}
    try:
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=';')
            for row in reader:
                if row.get('path'):
                    result[row['path']] = float(row.get('timestamp', 0))
    except Exception as e:
        log(f"ERREUR lecture dictionnaire : {e}", 1)
    return result


def get_next_archive_name(archives_dir: str, rel_path: str) -> str:
    """Génère le prochain nom d'archive pour un fichier"""
    ensure_directory(archives_dir)
    basename = os.path.basename(rel_path)
    name, ext = os.path.splitext(basename)
    existing = [f for f in os.listdir(archives_dir) if f.startswith(name) and f.endswith(ext)]
    numbers = []
    for f in existing:
        match = re.search(rf'{re.escape(name)}(\d+){re.escape(ext)}', f)
        if match:
            numbers.append(int(match.group(1)))
    next_num = max(numbers) + 1 if numbers else 1
    return os.path.join(archives_dir, f"{name}{next_num}{ext}")


def sauvegarde_cible(cible: CibleConfig):
    """Effectue la sauvegarde incrémentale d'une cible"""
    if not cible.sauve or not cible.actif:
        return
    
    log(f"🔍 Sauvegarde {cible.nom}...", 1)
    
    ensure_directory(cible.sauve)
    ensure_directory(cible.archives)
    
    dict_file = os.path.join(cible.sauve, "dictionnaire.csv")
    
    projet_dict = build_file_dict(cible.chemin)
    sauve_dict = load_dict_from_csv(dict_file)
    
    changes = 0
    for rel_path, ts in projet_dict.items():
        if rel_path not in sauve_dict or ts > sauve_dict.get(rel_path, 0):
            src = os.path.join(cible.chemin, rel_path)
            dst = os.path.join(cible.sauve, rel_path)
            
            if os.path.exists(dst):
                archive_path = get_next_archive_name(cible.archives, rel_path)
                try:
                    shutil.copy2(dst, archive_path)
                    log(f"📦 Archivé : {os.path.basename(archive_path)}", 2)
                except Exception as e:
                    log(f"ERREUR archivage : {e}", 2)
            
            try:
                os.makedirs(os.path.dirname(dst), exist_ok=True)
                shutil.copy2(src, dst)
                emoji = '🔄' if rel_path in sauve_dict else '✨'
                log(f"{emoji} {rel_path}", 2)
                sauve_dict[rel_path] = ts
                changes += 1
            except Exception as e:
                log(f"ERREUR copie : {e}", 2)
    
    if changes == 0:
        log(f"✓ Aucun changement ({cible.nom})", 1)
    else:
        log(f"✓ {changes} fichier(s) sauvegardé(s) ({cible.nom})", 1)
        save_dict_to_csv(sauve_dict, dict_file)


# ═══════════════════════════════════════════════════════════════════
# BACKUP SSD
# ═══════════════════════════════════════════════════════════════════

def backup_ssd():
    """Effectue le backup quotidien sur SSD si disponible"""
    if not os.path.exists(SSD_BASE_DIR):
        return
    
    log("✓ SSD accessible", 1)
    today = date_today()
    backup_dir = os.path.join(SSD_BASE_DIR, today)
    
    if os.path.exists(backup_dir):
        log(f"✓ Backup du jour déjà fait ({today})", 1)
        return
    
    log(f"💿 Backup SSD : {today}", 1)
    
    # Backup de toutes les cibles actives
    for pe in POINTS_ENTREE:
        for cible in pe.cibles:
            if cible.actif:
                cible_backup_dir = os.path.join(backup_dir, cible.nom)
                try:
                    shutil.copytree(cible.chemin, cible_backup_dir)
                    count = sum([len(files) for _, _, files in os.walk(cible_backup_dir)])
                    log(f"✓ {cible.nom} : {count} fichiers copiés", 2)
                except Exception as e:
                    log(f"ERREUR backup SSD {cible.nom} : {e}", 2)


# ═══════════════════════════════════════════════════════════════════
# INITIALISATION
# ═══════════════════════════════════════════════════════════════════

def creer_structure_initiale():
    """Crée la structure initiale des répertoires si nécessaire"""
    # Créer les points d'entrée
    for pe_dir in [CERBERE_TESTS_DIR, CERBERE_PROD_DIR, CERBERE_ALL_DIR]:
        ensure_directory(pe_dir)
    
    # Créer les fichiers de config par défaut s'ils n'existent pas
    for pe_dir, nom, chemin, sauve, archives in [
        (CERBERE_TESTS_DIR, "tests", r"c:\cx", r"c:\cxsauve", r"c:\cxarchives"),
        (CERBERE_PROD_DIR, "prod", r"c:\kitviewsearchV5", r"c:\kitviewsearchV5sauve", r"c:\kitviewsearchV5archives"),
    ]:
        config_file = os.path.join(pe_dir, "cerbere_config.csv")
        if not os.path.exists(config_file):
            log(f"📝 Création config par défaut : {config_file}", 1)
            with open(config_file, 'w', encoding='utf-8-sig', newline='') as f:
                f.write(f"# Configuration Cerbère pour environnement {nom.upper()}\n")
                f.write(f"# Généré automatiquement le {timestamp_now()}\n")
                f.write("#\n")
                f.write("type;parametre;valeur\n")
                f.write(f"cible;nom;{nom}\n")
                f.write(f"cible;chemin;{chemin}\n")
                f.write(f"cible;sauve;{sauve}\n")
                f.write(f"cible;archives;{archives}\n")
                f.write("cible;actif;oui\n")
                f.write("#\n")
                f.write("# Règles de dispatch (ordre = priorité)\n")
                f.write("regle;Prompt_*;.md;prompts;Prompts de projet\n")
                f.write("regle;prompt_*;.md;prompts;Prompts (minuscules)\n")
                f.write("regle;con*;.md;convs;Conversations\n")
                f.write("regle;conv_*;.md;convs;Conversations (underscore)\n")
                f.write("regle;*;.md;doc;Markdown par défaut\n")
                f.write("regle;logrecherche*;.csv;logs;Logs de recherche\n")
                f.write("regle;data*;.csv;data;Données\n")
                f.write("regle;*pat*;.csv;data;Fichiers patients\n")
                f.write("regle;test*;.csv;tests;Fichiers de test\n")
                f.write("regle;logs*;.csv;logs;Autres logs\n")
                f.write("regle;*;.csv;refs;CSV par défaut\n")
                f.write("regle;*;.py;;Scripts Python (racine)\n")
                f.write("regle;*;.html;ihm;Pages web\n")
                f.write("regle;*;.css;ihm;Feuilles de style\n")
                f.write("regle;*;.js;ihm;Scripts JavaScript\n")
                f.write("regle;*;.db;db;Bases de données\n")
                f.write("regle;*;.xlsx;doc;Documents Excel\n")
                f.write("regle;*;.docx;doc;Documents Word\n")
                f.write("regle;*;.txt;doc;Documents texte\n")
                f.write("regle;*;.pdf;doc;Documents PDF\n")
                f.write("regle;*;*;ztri;Fichiers non reconnus\n")


def initialiser_cibles():
    """Crée les sous-répertoires nécessaires pour chaque cible"""
    sous_repertoires = ['ihm', 'db', 'prompts', 'doc', 'convs', 'ztri', 'data', 'tests', 'logs', 'refs']
    
    for pe in POINTS_ENTREE:
        for cible in pe.cibles:
            if cible.actif:
                # Créer le répertoire cible principal
                ensure_directory(cible.chemin)
                
                # Créer les sous-répertoires
                for sous_rep in sous_repertoires:
                    ensure_directory(os.path.join(cible.chemin, sous_rep))
                
                # Créer les répertoires de sauvegarde
                if cible.sauve:
                    ensure_directory(cible.sauve)
                if cible.archives:
                    ensure_directory(cible.archives)


# ═══════════════════════════════════════════════════════════════════
# BOUCLE PRINCIPALE
# ═══════════════════════════════════════════════════════════════════

def main():
    global POINTS_ENTREE
    
    print(f"{__pgm__} V{__version__} - {__date__}\n")
    afficher_logo()
    
    # Vérifier openpyxl
    if not OPENPYXL_AVAILABLE:
        log("⚠️  ATTENTION : openpyxl non installé")
        log("   La conversion xlsx → csv ne fonctionnera pas", 1)
        log("   Installer avec : pip install openpyxl", 1)
    
    # Créer la structure initiale
    log("\n📂 Création structure initiale...")
    creer_structure_initiale()
    
    # Charger la configuration
    log("\n📋 Chargement configuration...")
    POINTS_ENTREE = charger_configuration()
    
    if not POINTS_ENTREE:
        log("❌ Aucun point d'entrée configuré !")
        log("   Vérifier les fichiers cerbere_config.csv dans :")
        log(f"   - {CERBERE_TESTS_DIR}", 1)
        log(f"   - {CERBERE_PROD_DIR}", 1)
        sys.exit(1)
    
    # Afficher la configuration
    log("\n📊 Configuration :")
    for pe in POINTS_ENTREE:
        log(f"Point d'entrée : {pe.chemin} ({pe.nom})", 1)
        for cible in pe.cibles:
            log(f"  → Cible : {cible.nom} ({cible.chemin})", 2)
            log(f"     Sauve : {cible.sauve}", 2)
            log(f"     Règles : {len(cible.regles)}", 2)
    
    # Initialiser les répertoires des cibles
    log("\n✓ Vérification des répertoires...")
    initialiser_cibles()
    
    # Initialiser les fichiers de version
    log("\n✓ Initialisation versions...")
    for pe in POINTS_ENTREE:
        init_version_files(pe)
    
    # Sauvegarde initiale
    log("\n✓ Sauvegarde initiale...")
    cibles_traitees = set()
    for pe in POINTS_ENTREE:
        for cible in pe.cibles:
            if cible.nom not in cibles_traitees:
                sauvegarde_cible(cible)
                cibles_traitees.add(cible.nom)
    
    log("\n🔄 Surveillance active (Ctrl+C pour arrêter)\n")
    
    try:
        while True:
            log(f"⏰ {timestamp_now()}\n")
            
            # Import pour chaque point d'entrée
            for pe in POINTS_ENTREE:
                log(f"🔄 IMPORT [{pe.nom.upper()}]")
                import_files(pe)
            
            log(f"\n⏸️  {CYCLE_WAIT}s...\n")
            time.sleep(CYCLE_WAIT)
            
            # Sauvegarde de chaque cible (sans doublon)
            log("💾 SAUVEGARDE")
            cibles_traitees = set()
            for pe in POINTS_ENTREE:
                for cible in pe.cibles:
                    if cible.nom not in cibles_traitees:
                        sauvegarde_cible(cible)
                        cibles_traitees.add(cible.nom)
            
            log(f"\n⏸️  {CYCLE_WAIT}s...\n")
            time.sleep(CYCLE_WAIT)
            
            # Backup SSD
            log("💿 BACKUP SSD")
            backup_ssd()
            
            log("\n" + "=" * 60 + "\n")
    
    except KeyboardInterrupt:
        log("\n\n🛑 Arrêt demandé")
        log("✓ Cerbère se rendort...\n")
        sys.exit(0)


if __name__ == "__main__":
    main()
